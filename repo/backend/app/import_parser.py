"""Excel upload parser for the driver routing backend prototype.

This module depends on `openpyxl` to read `.xlsx` workbooks and normalize
worksheet rows into the dictionaries expected by `RoutingService.import_orders_from_rows`.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def _header_map(headers: list[str]) -> dict[str, int]:
    normalized = {str(header).strip().lower(): idx for idx, header in enumerate(headers)}
    return normalized


def _cell_value(cell: Any) -> object:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def parse_xlsx_rows(file_bytes: bytes, sheet_name: str | None = None) -> list[dict[str, object]]:
    """Parse the first worksheet of an uploaded `.xlsx` file into row dicts.

    The first non-empty row is treated as the header. Header cells are
    normalized to lowercase and stripped before matching against the
    `RoutingService.import_orders_from_rows` expected fields.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    rows: list[dict[str, object]] = []
    header: list[str] = []
    header_map: dict[str, int] = {}
    for row in ws.iter_rows(values_only=False):
        raw_values = [_cell_value(cell) for cell in row]
        if not any(value not in (None, "") for value in raw_values):
            continue
        if not header:
            header = [str(value).strip() if value is not None else "" for value in raw_values]
            header_map = _header_map(header)
            continue
        row_dict: dict[str, object] = {}
        for field_name, idx in header_map.items():
            if idx < len(raw_values):
                row_dict[field_name] = raw_values[idx]
        rows.append(row_dict)

    wb.close()
    return rows
